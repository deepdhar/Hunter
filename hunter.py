import os
import base64
import email
import email.encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import mimetypes
import os
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from tkinter import *
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random, string
from bs4 import BeautifulSoup
from fpdf import FPDF
import pdfkit
import imgkit
import time
import threading
import pyrebase
import firebase_admin
from firebase_admin import db, credentials


SCOPES = ["https://mail.google.com/"]

root = Tk()
root.configure(bg='#660dff')
root.title("Hunter")
root.geometry("640x400")

bodyIndex=1
subIndex=1
nameIndex=1
currentSenderEmail=''
currentReceiverEmail=''
currentBody=''
currentSubject=''
currentSenderName=''
mailingTo = currentSenderEmail
limitCount = 0
fixEmailCount = 0

randomNum = 0
randomInvoice = 0
randomTransaction = 0
randomItem = 0
currentEmailCount = 0 #used in startSendingEmail function
currentSenderCountInput = 0 #used in senderButtonPressed function
namecodeFlag = FALSE

stopThreads = False;

def Home(username):
    global NewRoot
    global limitCount
    global fixEmailCount
    
    root.withdraw() # hide (close) the root/Tk window
    NewRoot = Toplevel(root)
    NewRoot.title("Hunter")
    NewRoot.geometry("1200x640")
    # use the NewRoot as the root now
    
    cred = credentials.Certificate("firebase_credentials.json")
    firebase_admin.initialize_app(cred, {"databaseURL":"https://hunter-enterprise-default-rtdb.asia-southeast1.firebasedatabase.app/"})
    user_db_path = '/' + username
    limitCount = db.reference(user_db_path + '/dailyLimit').get()
    renewal_date = db.reference(user_db_path + '/renewalDate').get()

    totalFromSender = 1
    
    def getSubject():
        path = "subjects.xlsx"
        workbook = load_workbook(path)
        worksheet = workbook.active
        length = len(list(worksheet.values))
        global subIndex
        subIndex = random.randint(1, length)
        subjectInput.delete(0,END)
        new_subject = worksheet["A"+str(subIndex)].value
        subjectInput.insert(0,new_subject)
        
    def getSenderName():
        path = "names.xlsx"
        workbook = load_workbook(path)
        worksheet = workbook.active
        length = len(list(worksheet.values))
        global nameIndex
        nameIndex = random.randint(1, length)
        senderNameInput.delete(0,END)
        new_name = worksheet["A"+str(nameIndex)].value
        if namecodeFlag:
            new_name = new_name + "$NAMECODE"
            senderNameInput.insert(0,new_name)
        else:
            senderNameInput.insert(0,new_name)

    def loadSenders():
        path = "senders.xlsx"
        workbook = load_workbook(path)
        worksheet = workbook.active
        senderEmailInput.delete(0,END)
        
        length = len(list(worksheet.values))
        for i in range(1, length+1):
            # print(worksheet["B"+str(i)].value)
            buttons = [senderEmailButton1, senderEmailButton2, senderEmailButton3, senderEmailButton4, senderEmailButton5, senderEmailButton6, senderEmailButton7, senderEmailButton8, senderEmailButton9, senderEmailButton10, senderEmailButton11, senderEmailButton12, senderEmailButton13, senderEmailButton14, senderEmailButton15]
            buttons[i-1].configure(text=worksheet["A"+str(i)].value, background='#b1e6fc')

    def senderButtonPressed(button):
        global currentSenderEmail
        global currentSenderCountInput
        global currentEmailCount
        buttonText = button.cget('text')
        buttons = [senderEmailButton1, senderEmailButton2, senderEmailButton3, senderEmailButton4, senderEmailButton5, senderEmailButton6, senderEmailButton7, senderEmailButton8, senderEmailButton9, senderEmailButton10, senderEmailButton11, senderEmailButton12, senderEmailButton13, senderEmailButton14, senderEmailButton15]
        buttonCountInputs = [senderEmailCount1, senderEmailCount2, senderEmailCount3, senderEmailCount4, senderEmailCount5, senderEmailCount6, senderEmailCount7, senderEmailCount8, senderEmailCount9, senderEmailCount10, senderEmailCount11, senderEmailCount12, senderEmailCount13, senderEmailCount14, senderEmailCount15]
        
        senderEmailCountInput.delete(0,END)
        currentEmailCount = 0
        
        for i in range(1, len(buttons)):
            if button==buttons[i-1]:
                currentSenderCountInput = buttonCountInputs[i-1]
                # currentSenderCountInput.get(0,END)
                # pass current sender email count input value to sender email count input at top
                temp = currentSenderCountInput.get()
                # currentEmailCount = int(float(temp))
                print(temp)
                senderEmailCountInput.insert(0, temp)
                break 
        
        if "Not Updated" in buttonText:
            return
        else:
            currentSenderEmail = buttonText
            senderEmailInput.delete(0,END)
            senderEmailInput.insert(0,currentSenderEmail)


    def loadBody():
        path = "bodys.xlsx"
        workbook = load_workbook(path)
        worksheet = workbook.active
        length = len(list(worksheet.values))
        global bodyIndex
        bodyIndex = random.randint(1, length)
        bodyInput.delete('1.0', END)
        new_body = worksheet["A"+str(bodyIndex)].value
        bodyInput.insert('1.0', new_body)
        senderEmailCount1.delete(0,END)
        senderEmailCount1.insert(0,'1')

    def getEmailService(credentials):
        flow = InstalledAppFlow.from_client_secrets_file(credentials+".json", SCOPES)
        creds = flow.run_local_server(port=0)
        service = build("gmail", "v1", credentials=creds)
        return service

    def create_message_with_attachment(sender, to, subject, displayName, message_text, file):
        message = MIMEMultipart('mixed')
        message['to'] = to
        message['from'] = f'{displayName} <{sender}>'
        message['subject'] = subject
        message['Reply-To'] = sender
        
        text_part = MIMEText(message_text, 'plain')
        message.attach(text_part)
        
        # msg_html_part = MIMEText(message_text, 'html')
        # message.attach(msg_html_part)
        
       # Attach the PDF file
        if file:
            content_type, encoding = mimetypes.guess_type(file)
            if content_type is None or encoding is not None:
                content_type = 'application/octet-stream'
            
            main_type, sub_type = content_type.split('/', 1)
            with open(file, 'rb') as pdf_file:
                mime_part = MIMEBase(main_type, sub_type)
                mime_part.set_payload(pdf_file.read())
                email.encoders.encode_base64(mime_part)
                mime_part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file)}"')
                message.attach(mime_part)
        
        raw = base64.urlsafe_b64encode(message.as_bytes())
        raw = raw.decode()
        return {'raw': raw}

    def create_message_with_attachment2(sender, to, subject, displayName, message_text, file):
        message = MIMEMultipart()
        message['to'] = to
        message['from'] = f'{displayName} <{sender}>'
        message['subject'] = subject
        msg = MIMEText(message_text, 'html')
        message.attach(msg)
        content_type, encoding = mimetypes.guess_type(file)
        main_type, sub_type = content_type.split('/', 1)
        fp = open(file, 'rb')
        msg = MIMEBase(main_type, sub_type)
        msg.set_payload(fp.read())
        fp.close()
        # filename = os.path.basename(file)
        msg.add_header('Content-Disposition', 'attachment', filename=file)
        email.encoders.encode_base64(msg)
        message.attach(msg)
        raw_message = base64.urlsafe_b64encode(message.as_string().encode("utf-8"))
        return {'raw': raw_message.decode("utf-8")}
        
    def send_message(service, user_id, message):
        try:
          message = service.users().messages().send(userId=user_id, body=message).execute()
          print('Message Id: %s' % message['id'])
          return message
        except Exception as e:
          print('An error occurred: %s' % e)
          return None
        
    def sendEmail(sender, to, subject, displayName, message_text, service, file):
        messageContent = create_message_with_attachment(sender, to, subject, displayName, message_text, file)
        messageId = send_message(service, user_id="me", message=messageContent)
        return messageId
 
    def callStart():
        global stopThreads
        if stopThreads==False:
            t1 = threading.Thread(target=startSendingEmail)
            print("Thread running: " + str(t1))
            t1.start()
        # else:
        #     t2 = threading.Thread(target=startSendingEmail)
        #     t1.join(t2)
        #     print("Prev thread stopped")
        #     stopThreads = False

    def getRandomNum():
        global randomNum
        
        temp1 = ''.join(random.choices(string.ascii_uppercase, k=3)) + "-"
        temp2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp3 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp4 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-"
        temp5 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=11))
        
        randomNum = temp1+temp2+temp3+temp4+temp5
        
        return randomNum
    
    def getRandomInvoice():
        global randomInvoice
        
        temp1 = ''.join(random.choices(string.ascii_uppercase, k=3)) + "-"
        temp2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp3 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp4 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-"
        temp5 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=11))
        
        randomInvoice = temp1+temp2+temp3+temp4+temp5
        
        return randomInvoice
    
    def getRandomTransaction():
        global randomTransaction
        
        temp1 = ''.join(random.choices(string.ascii_uppercase, k=3)) + "-"
        temp2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp3 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp4 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-"
        temp5 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=11))
        
        randomTransaction = temp1+temp2+temp3+temp4+temp5
        
        return randomTransaction
    
    def getRandomItem():
        global randomItem
        
        temp1 = ''.join(random.choices(string.ascii_uppercase, k=3)) + "-"
        temp2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp3 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3)) + "-"
        temp4 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-"
        temp5 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=11))
        
        randomItem = temp1+temp2+temp3+temp4+temp5
        
        return randomItem

    def getRandomNameCode():
        temp = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
        return '[' + temp + ']';

    def startSendingEmail():
        global limitCount
        if senderEmailInput.get()=='' or subjectInput.get()=='' or senderNameInput.get()=='' or len(bodyInput.get('1.0', 'end-1c'))==0 or len(htmlInput.get('1.0', 'end-1c'))==0 or len(receiversInput.get('1.0', 'end-1c'))==0:
            print('check empty values')
        elif limitCount==0:
            messagebox.showerror(title='Error', message="Your daily limit is expired. Contact the admin.")
        else:
            global currentEmailCount
            global currentSenderCountInput
            global namecodeFlag
            fixEmailCount = int(fixEmailCountInput.get())
            service = getEmailService("client_secret_"+senderEmailInput.get())
            namecodeFlag = FALSE
            while len(receiversInput.get('1.0', 'end-1c'))!=0 and currentEmailCount<fixEmailCount:
                if limitCount==0:
                    messagebox.showerror(title='Error', message="Your daily limit is expired. Contact the admin.")
                    return
                global currentReceiverEmail
                global currentBody
                global currentSubject
                global currentSenderName
                global namecodeFlag
                
                currentReceiverEmail = receiversInput.get('1.0','2.0');
                currentReceiverEmail = currentReceiverEmail.strip()
                print(currentReceiverEmail)
                mailToLabel.config(text="Mailing to: " + str(currentReceiverEmail))
                
                currentSubject = subjectInput.get()
                currentBody = bodyInput.get('1.0', END)
                currentSenderName = senderNameInput.get()
                if "$NAMECODE$" in currentSenderName:
                    namecodeFlag = TRUE;
                currentSenderName = currentSenderName.replace("$NAMECODE$", str(getRandomNameCode()))
                getSenderName()
                # getSubject()
                # loadBody()

                
                currentSubject = currentSubject.replace("$RANDOM$", str(getRandomNum()))
                currentSubject = currentSubject.replace("$INVOICE$", str(getRandomInvoice()))
                currentSubject = currentSubject.replace("$TRANSACTION$", str(getRandomTransaction()))
                currentSubject = currentSubject.replace("$ITEMNO$", str(getRandomItem()))
                
                currentBody = currentBody.replace("$RANDOM$", str(getRandomNum()))
                currentBody = currentBody.replace("$INVOICE$", str(getRandomInvoice()))
                currentBody = currentBody.replace("$TRANSACTION$", str(getRandomTransaction()))
                currentBody = currentBody.replace("$ITEMNO$", str(getRandomItem()))
                currentBody = currentBody.replace("$EMAIL$", currentReceiverEmail)

                html = htmlInput.get('1.0',END);
                # soup = BeautifulSoup(html,"html-parser")
                # saveToPDF(soup.get_text())
                pdfName = saveToPDF(html)
                
                sendEmail(senderEmailInput.get(), currentReceiverEmail, currentSubject, currentSenderName, currentBody, service, pdfName)
                temp = random.randint(1, 3)
                time.sleep(temp)
                
                
                currentEmailCount = currentEmailCount + 1
                # update in top sender count input
                senderEmailCountInput.delete(0,END)
                senderEmailCountInput.insert(0, currentEmailCount)
                
                # update in list sender count input
                currentSenderCountInput.delete(0,END)
                currentSenderCountInput.insert(0, currentEmailCount)
                
                receiversInput.delete('1.0','2.0');
                limitCount = limitCount - 1
                db.reference(user_db_path + '/dailyLimit').set(limitCount)
                remainingLimitLabel.config(text = "Remaining limit: "+str(limitCount))
                
                # time.sleep(1)

        # time.sleep(1)
                
    def saveToPDF(htmlText):
        global currentReceiverEmail
        
        htmlText = htmlText.replace("$RANDOM$", str(getRandomNum()))
        htmlText = htmlText.replace("$INVOICE$", str(getRandomInvoice()))
        htmlText = htmlText.replace("$TRANSACTION$", str(getRandomTransaction()))
        htmlText = htmlText.replace("$ITEMNO$", str(getRandomItem()))
        htmlText = htmlText.replace("$EMAIL$", currentReceiverEmail)

        path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        pdf_filename = str(getRandomNum())  + ".pdf"

        pdfkit.from_string(htmlText, pdf_filename, configuration=config)
        
        return pdf_filename
      
    def loadReceivers():
        path = "receivers.xlsx"
        workbook = load_workbook(path)
        worksheet = workbook.active
        length = len(list(worksheet.values))
        receiversInput.delete('1.0', END)
        for i in range(1, length+1):
            # text=worksheet["B"+str(i)].value
            receiverEmail = worksheet["A"+str(i)].value
            print(receiverEmail)
            receiverInputIndex = str(i+1)
            receiverInputIndex = receiverInputIndex
            receiversInput.insert('1.0', receiverEmail + "\n")


    # first column section
    getSubjectButton = Button(NewRoot, text="Get Subject", padx=20, pady=0, background="#b1e6fc", font=('Arial 10'), anchor='center', command=getSubject)
    getSubjectButton.grid(row=0, column=0, columnspan=2, ipady=4)

    subjectInput = Entry(NewRoot, width=48, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    subjectInput.grid(row=1, column=0, columnspan=2, padx=15, pady=10, ipady=4)


    # second column section
    senderEmaillabel = Label(NewRoot, text="Sender Email: ", font=('Arial, 11'), anchor="w",)
    senderEmaillabel.grid(row=0, column=2, columnspan=2,)

    senderEmailInput = Entry(NewRoot, width=48, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailInput.grid(row=1, column=2, columnspan=2, padx=15, pady=10, ipady=4, ipadx=5)
    senderEmailCountInput = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCountInput.grid(row=1, column=4, ipady=4)


    # third column section
    senderNameInput = Entry(NewRoot, width=48, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderNameInput.grid(row=1, column=5, columnspan=2, padx=15, pady=10, ipady=4)

    # autoSenderNameButton = Button(NewRoot, text='Auto Load Name', background='#ff4255', anchor='w')
    # autoSenderNameButton.grid(row=0, column=5)

    getSenderNameButton = Button(NewRoot, text='Get Sender Name', background='#b1e6fc', anchor='e', command=getSenderName)
    getSenderNameButton.grid(row=0, column=6, padx=(0,15))


    # third row (row=2)
    loadSendersButton = Button(NewRoot, text='Load Senders', background='#b1e6fc', command=loadSenders)
    loadSendersButton.grid(row=2, column=0, pady=15)

    chooseJsonFolderButton = Button(NewRoot, text='Choose JSON Folder', background='#b1e6fc')
    chooseJsonFolderButton.grid(row=2, column=1, pady=15, padx=(0,15))


    mailToLabel = Label(NewRoot, text="Start mailing!", font=('Arial, 11'), anchor="w")
    mailToLabel.grid(row=2, column=2, columnspan=2)

    remainingLimitLabel = Label(NewRoot, text="Remaining Limit: " + str(limitCount), font=('Arial, 11'), anchor="w")
    remainingLimitLabel.grid(row=2, column=5, columnspan=1)


    # fourth row (row=3)
    # first column buttons
    senderEmailButton1 = Button(NewRoot, text="1. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton1))
    senderEmailButton1.grid(row=3, column=0, padx=(5,0))
    senderEmailCount1 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount1.grid(row=3, column=1)

    senderEmailButton2 = Button(NewRoot, text="2. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton2))
    senderEmailButton2.grid(row=4, column=0, padx=(5,0))
    senderEmailCount2 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount2.grid(row=4, column=1)

    senderEmailButton3 = Button(NewRoot, text="3. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton3))
    senderEmailButton3.grid(row=5, column=0, padx=(5,0))
    senderEmailCount3 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount3.grid(row=5, column=1)

    senderEmailButton4 = Button(NewRoot, text="4. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton4))
    senderEmailButton4.grid(row=6, column=0, padx=(5,0))
    senderEmailCount4 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount4.grid(row=6, column=1)

    senderEmailButton5 = Button(NewRoot, text="5. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton5))
    senderEmailButton5.grid(row=7, column=0, padx=(5,0))
    senderEmailCount5 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount5.grid(row=7, column=1)

    # third column buttons
    senderEmailButton6 = Button(NewRoot, text="6. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton6))
    senderEmailButton6.grid(row=3, column=2, padx=(5,0))
    senderEmailCount6 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount6.grid(row=3, column=3)

    senderEmailButton7 = Button(NewRoot, text="7. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton7))
    senderEmailButton7.grid(row=4, column=2, padx=(5,0))
    senderEmailCount7 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount7.grid(row=4, column=3)

    senderEmailButton8 = Button(NewRoot, text="8. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton8))
    senderEmailButton8.grid(row=5, column=2, padx=(5,0))
    senderEmailCount8 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount8.grid(row=5, column=3)

    senderEmailButton9 = Button(NewRoot, text="9. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton9))
    senderEmailButton9.grid(row=6, column=2, padx=(5,0))
    senderEmailCount9 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount9.grid(row=6, column=3)

    senderEmailButton10 = Button(NewRoot, text="10. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton10))
    senderEmailButton10.grid(row=7, column=2, padx=(5,0))
    senderEmailCount10 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount10.grid(row=7, column=3)

    # fifth column buttons
    senderEmailButton11 = Button(NewRoot, text="11. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton11))
    senderEmailButton11.grid(row=3, column=5, padx=(5,0))
    senderEmailCount11 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount11.grid(row=3, column=6)

    senderEmailButton12 = Button(NewRoot, text="12. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton12))
    senderEmailButton12.grid(row=4, column=5, padx=(5,0))
    senderEmailCount12 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount12.grid(row=4, column=6)

    senderEmailButton13 = Button(NewRoot, text="13. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton13))
    senderEmailButton13.grid(row=5, column=5, padx=(5,0))
    senderEmailCount13 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount13.grid(row=5, column=6)

    senderEmailButton14 = Button(NewRoot, text="14. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton14))
    senderEmailButton14.grid(row=6, column=5, padx=(5,0))
    senderEmailCount14 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount14.grid(row=6, column=6)

    senderEmailButton15 = Button(NewRoot, text="15. Not Updated", background='#bf82ed', width=30, anchor='w', padx=5, command=lambda: senderButtonPressed(senderEmailButton15))
    senderEmailButton15.grid(row=7, column=5, padx=(5,0))
    senderEmailCount15 = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    senderEmailCount15.grid(row=7, column=6)


    # seventh row (row=8)
    tagsTextBox = Text(NewRoot, width=45, borderwidth=5, height=3, background='#90f5e6')
    tagsTextBox.grid(row=8, column=2, columnspan=3, pady=(20,10))
    tagsTextBox.insert('1.0', "Tags:- $INVOICE$, $TRANSACTION$, $DATE$, \n $EMAIL$, $ITEMNO$, $RANDOM$", "$NAMECODE$")
    tagsTextBox.config(state=DISABLED)

    renewalLimitLabel = Label(NewRoot, text="next renewal date: " + str(renewal_date), font=('Arial, 11'), anchor="w")
    renewalLimitLabel.grid(row=8, column=5, columnspan=1, pady=20)
    
    # everything starts from this button
    startButton = Button(NewRoot, text="Start", background='#15d629', width=10, font=('Arial, 11'), command=callStart)
    startButton.grid(row=8, column=6, pady=20)


    # eigth row (row=9)
    htmlLabel = Label(NewRoot, text="Paste Html for PDF:", font=('Arial, 10'), anchor='center')
    htmlLabel.grid(row=9, column=0, columnspan=2, pady=10)
    htmlInput = Text(NewRoot, width=42, height=12, background='#90f5e6')
    htmlInput.grid(row=10, column=0, columnspan=2)

    loadBodyButton = Button(NewRoot, text='Load Body', background="#b1e6fc", font=('Arial 10'), anchor='center', command=loadBody)
    loadBodyButton.grid(row=9, column=2, columnspan=3)
    bodyInput = Text(NewRoot, width=45, height=12, background='#90f5e6')
    bodyInput.grid(row=10, column=2, columnspan=3)

    loadReceiversButton = Button(NewRoot, text='Load Receivers', background="#b1e6fc", font=('Arial 10'), anchor='center', command=loadReceivers)
    loadReceiversButton.grid(row=9, column=5, columnspan=1)
    fixEmailCountInput = Entry(NewRoot, width=4, borderwidth=5, font=('Arial 10'), background="#90f5e6")
    fixEmailCountInput.grid(row=9, column=6)
    fixEmailCountInput.insert(0, 5)
    receiversInput = Text(NewRoot, width=45, height=12, background='#90f5e6')
    receiversInput.grid(row=10, column=5, columnspan=3)
    

firebaseConfig = {    
    'apiKey': "AIzaSyDsP2FA_EM7NqPuLYZZKKUdwvUQcdczfgQ",
    'authDomain': "hunter-enterprise.firebaseapp.com",
    'databaseURL': "https://trialauth-7eea1.firebaseio.com",
    'projectId': "hunter-enterprise",
    'storageBucket': "hunter-enterprise.appspot.com",
    'messagingSenderId': "932657854963",
    'appId': "1:932657854963:web:dbd30a1495facd5200ba86",
    'measurementId': "G-7NFEEXWW65"
}

firebase = pyrebase.initialize_app(firebaseConfig)
auth = firebase.auth()

def login():
    email = username_entry.get()
    password = password_entry.get()
    
    try:
        auth.sign_in_with_email_and_password(email, password)
        username = email.replace("@gmail.com", "")
        root.after(1000, Home(username))
    except:
        messagebox.showerror(title='Error', message="Invalid login.")
    return
    

    
frame = Frame(bg='#660dff')

login_label = Label(frame, text="Hunter", bg='#660dff', fg="#FFFFFF", font=("Arial", 30))
login_label.grid(row=0, column=0, columnspan=2, sticky="news", pady=40)

username_label = Label(frame, text="Username", bg='#660dff', fg="#FFFFFF", font=("Arial", 16, 'bold'))
username_label.grid(row=1, column=0, padx=(0,20))

password_label = Label(frame, text="Password", bg='#660dff', fg="#FFFFFF", font=("Arial", 16, 'bold'))
password_label.grid(row=2, column=0, padx=(0,20))

username_entry = Entry(frame, font=("Arial", 16))
username_entry.grid(row=1, column=1, pady=20)

password_entry = Entry(frame, show="*", font=("Arial", 16))
password_entry.grid(row=2, column=1, pady=20)

login_button = Button(frame, text="Sign In", bg="#DC143C", fg="#FFFFFF", padx=20, pady=1, font=("Arial", 16), command=login)
login_button.grid(row=3, column=0, columnspan=2, pady=20)



frame.pack()
root.mainloop()